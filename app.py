from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from google_sheets_db import GoogleSheetsDB
from config import Config
from datetime import datetime
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
try:
    import barcode
    from barcode.writer import ImageWriter
    BARCODE_AVAILABLE = True
except ImportError:
    BARCODE_AVAILABLE = False
from PIL import Image

app = Flask(__name__)
app.config.from_object(Config)

# Configure upload folders
UPLOAD_FOLDER = 'uploads'
IMAGE_FOLDER = os.path.join(UPLOAD_FOLDER, 'images')
DOCUMENT_FOLDER = os.path.join(UPLOAD_FOLDER, 'documents')
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
ALLOWED_DOCUMENT_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt'}

# Create upload directories if they don't exist
os.makedirs(IMAGE_FOLDER, exist_ok=True)
os.makedirs(DOCUMENT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

def allowed_file(filename, allowed_extensions):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

# Initialize database
try:
    db = GoogleSheetsDB()
    print("[SUCCESS] Google Sheets connection established successfully")
except FileNotFoundError as e:
    print(f"[ERROR] Credentials file not found: {e}")
    print("Please ensure credentials.json exists in the project root directory.")
    db = None
except Exception as e:
    print(f"[ERROR] Could not initialize Google Sheets connection: {e}")
    print("Please check:")
    print("  1. credentials.json exists and is valid")
    print("  2. The service account email has been shared with the Google Sheet")
    print("  3. Google Sheets API is enabled in your Google Cloud project")
    import traceback
    traceback.print_exc()
    db = None

def log_activity(action, entity_type, entity_id, description, details=''):
    """Helper function to log activities"""
    if not db:
        return
    try:
        log_id = db.get_next_id('ActivityLogs')
        log_data = {
            'ID': log_id,
            'Date & Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Type': 'Activity',
            'User': session.get('user_id', 'Unknown'),
            'Action': action,
            'Entity Type': entity_type,
            'Entity ID': str(entity_id),
            'Description': description,
            'Details': details
        }
        db.insert('ActivityLogs', log_data)
    except Exception as e:
        print(f"Error logging activity: {e}")

def login_required(f):
    """Decorator to require login"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to require admin role"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Access denied. Admin privileges required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if not db:
        flash('Database not configured. Please check your Google Sheets setup.', 'danger')
        return render_template('register.html')
        
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role', 'user')
        
        # Check if first user (make admin)
        users = db.get_all('Users')
        if not users:
            role = 'admin'
        
        # Check if username exists
        if not db:
            flash('Database not configured', 'danger')
            return render_template('register.html')
        
        existing_user = db.get_by_id('Users', 'Username', username)
        if existing_user:
            flash('Username already exists', 'danger')
            return render_template('register.html')
        
        # Hash password
        hashed_password = generate_password_hash(password)
        
        # Insert user
        user_data = {
            'Username': username,
            'Email': email,
            'Password': hashed_password,
            'Role': role
        }
        
        if db.insert('Users', user_data):
            flash('Registration successful! Please login.', 'success')
            return redirect(url_for('login'))
        else:
            flash('Registration failed. Please try again.', 'danger')
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if not db:
            flash('Database not configured. Please check your Google Sheets setup.', 'danger')
            return render_template('login.html')
        
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = db.get_by_id('Users', 'Username', username)
        
        if user and check_password_hash(user.get('Password', ''), password):
            session['user_id'] = username
            session['role'] = user.get('Role', 'user')
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    if not db:
        flash('Database not configured. Please check your Google Sheets setup.', 'danger')
        return redirect(url_for('login'))
    
    assets = db.get_all('Assets')
    categories = db.get_all('Categories')
    locations = db.get_all('Locations')
    
    # Prepare data for graphs
    # Assets by Category
    category_data = {}
    for asset in assets:
        category = asset.get('Asset Category', 'Uncategorized')
        category_data[category] = category_data.get(category, 0) + 1
    
    # Assets by Location
    location_data = {}
    for asset in assets:
        location = asset.get('Location', 'No Location')
        location_data[location] = location_data.get(location, 0) + 1
    
    # Assets by Status
    status_data = {}
    for asset in assets:
        status = asset.get('Asset Status', 'No Status')
        status_data[status] = status_data.get(status, 0) + 1
    
    # Total asset value
    total_value = 0
    for asset in assets:
        try:
            amount = float(asset.get('Amount', 0) or 0)
            total_value += amount
        except (ValueError, TypeError):
            pass
    
    # Assets by Brand (top 10)
    brand_data = {}
    for asset in assets:
        brand = asset.get('Brand', 'No Brand')
        brand_data[brand] = brand_data.get(brand, 0) + 1
    top_brands = sorted(brand_data.items(), key=lambda x: x[1], reverse=True)[:10]
    
    return render_template('dashboard.html', 
                         assets=assets, 
                         role=session.get('role'),
                         category_data=category_data,
                         location_data=location_data,
                         status_data=status_data,
                         brand_data=dict(top_brands),
                         total_value=total_value)

# Master Data Routes
@app.route('/locations')
@login_required
def locations():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('dashboard'))
    locations = db.get_all('Locations')
    return render_template('locations.html', locations=locations, role=session.get('role'))

@app.route('/locations/add', methods=['POST'])
@login_required
def add_location():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('locations'))
    location_name = request.form.get('location_name')
    if location_name:
        location_id = db.get_next_id('Locations')
        db.insert('Locations', {'ID': location_id, 'Location Name': location_name})
        flash('Location added successfully', 'success')
    return redirect(url_for('locations'))

@app.route('/locations/delete/<int:location_id>', methods=['POST'])
@admin_required
def delete_location(location_id):
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('locations'))
    if db.delete('Locations', 'ID', location_id):
        flash('Location deleted successfully', 'success')
    return redirect(url_for('locations'))

@app.route('/categories')
@login_required
def categories():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('dashboard'))
    categories = db.get_all('Categories')
    return render_template('categories.html', categories=categories, role=session.get('role'))

@app.route('/categories/add', methods=['POST'])
@login_required
def add_category():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('categories'))
    category_name = request.form.get('category_name')
    if category_name:
        category_id = db.get_next_id('Categories')
        db.insert('Categories', {'ID': category_id, 'Category Name': category_name})
        flash('Category added successfully', 'success')
    return redirect(url_for('categories'))

@app.route('/categories/delete/<int:category_id>', methods=['POST'])
@admin_required
def delete_category(category_id):
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('categories'))
    if db.delete('Categories', 'ID', category_id):
        flash('Category deleted successfully', 'success')
    return redirect(url_for('categories'))

@app.route('/subcategories')
@login_required
def subcategories():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('dashboard'))
    subcategories = db.get_all('Subcategories')
    categories = db.get_all('Categories')
    # Map category IDs to names
    category_map = {str(cat.get('ID', '')): cat.get('Category Name', '') for cat in categories}
    for sub in subcategories:
        sub['Category Name'] = category_map.get(str(sub.get('Category ID', '')), '')
    return render_template('subcategories.html', subcategories=subcategories, 
                         categories=categories, role=session.get('role'))

@app.route('/subcategories/add', methods=['POST'])
@login_required
def add_subcategory():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('subcategories'))
    subcategory_name = request.form.get('subcategory_name')
    category_id = request.form.get('category_id')
    if subcategory_name and category_id:
        subcategory_id = db.get_next_id('Subcategories')
        db.insert('Subcategories', {
            'ID': subcategory_id,
            'Subcategory Name': subcategory_name,
            'Category ID': category_id
        })
        flash('Subcategory added successfully', 'success')
    return redirect(url_for('subcategories'))

@app.route('/subcategories/delete/<int:subcategory_id>', methods=['POST'])
@admin_required
def delete_subcategory(subcategory_id):
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('subcategories'))
    if db.delete('Subcategories', 'ID', subcategory_id):
        flash('Subcategory deleted successfully', 'success')
    return redirect(url_for('subcategories'))

@app.route('/asset_types')
@login_required
def asset_types():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('dashboard'))
    asset_types = db.get_all('AssetTypes')
    return render_template('asset_types.html', asset_types=asset_types, role=session.get('role'))

@app.route('/asset_types/add', methods=['POST'])
@login_required
def add_asset_type():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('asset_types'))
    asset_code = request.form.get('asset_code')
    asset_type = request.form.get('asset_type')
    depreciation_value = request.form.get('depreciation_value')
    if asset_code and asset_type and depreciation_value:
        db.insert('AssetTypes', {
            'Asset Code': asset_code,
            'Asset Type': asset_type,
            'Depreciation Value (%)': depreciation_value
        })
        flash('Asset Type added successfully', 'success')
    return redirect(url_for('asset_types'))

@app.route('/asset_types/delete/<asset_code>', methods=['POST'])
@admin_required
def delete_asset_type(asset_code):
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('asset_types'))
    if db.delete('AssetTypes', 'Asset Code', asset_code):
        flash('Asset Type deleted successfully', 'success')
    return redirect(url_for('asset_types'))

@app.route('/brands')
@login_required
def brands():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('dashboard'))
    brands = db.get_all('Brands')
    return render_template('brands.html', brands=brands, role=session.get('role'))

@app.route('/brands/add', methods=['POST'])
@login_required
def add_brand():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('brands'))
    brand_name = request.form.get('brand_name')
    if brand_name:
        brand_id = db.get_next_id('Brands')
        db.insert('Brands', {'ID': brand_id, 'Brand Name': brand_name})
        flash('Brand added successfully', 'success')
    return redirect(url_for('brands'))

@app.route('/brands/delete/<int:brand_id>', methods=['POST'])
@admin_required
def delete_brand(brand_id):
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('brands'))
    if db.delete('Brands', 'ID', brand_id):
        flash('Brand deleted successfully', 'success')
    return redirect(url_for('brands'))

# Asset Entry Routes
@app.route('/assets')
@login_required
def assets():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('dashboard'))
    
    assets_list = db.get_all('Assets')
    # Debug: Print number of assets retrieved
    print(f"DEBUG: Retrieved {len(assets_list)} assets from Google Sheets")
    if assets_list:
        print(f"DEBUG: First asset: {assets_list[0]}")
    return render_template('assets.html', assets=assets_list, role=session.get('role'))

@app.route('/assets/add', methods=['GET', 'POST'])
@login_required
def add_asset():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('assets'))
    
    if request.method == 'POST':
        # Get form data
        item_name = request.form.get('item_name')
        asset_category = request.form.get('asset_category')
        asset_subcategory = request.form.get('asset_subcategory')
        brand = request.form.get('brand')
        asset_description = request.form.get('asset_description')
        amount = request.form.get('amount')
        location = request.form.get('location')
        date_of_purchase = request.form.get('date_of_purchase')
        warranty = request.form.get('warranty')
        department = request.form.get('department')
        ownership = request.form.get('ownership')
        
        # Generate asset code (using timestamp + category for uniqueness)
        asset_type_records = db.get_all('AssetTypes')
        # Simple auto-generation: TIMESTAMP format
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        asset_code = f"AST-{timestamp}"
        
        # Handle image upload
        image_filename = ''
        if 'image_attachment' in request.files:
            image_file = request.files['image_attachment']
            if image_file and image_file.filename and allowed_file(image_file.filename, ALLOWED_IMAGE_EXTENSIONS):
                filename = secure_filename(image_file.filename)
                # Add asset code to filename for uniqueness
                name, ext = os.path.splitext(filename)
                image_filename = f"{asset_code}_{name}{ext}"
                image_path = os.path.join(IMAGE_FOLDER, image_filename)
                image_file.save(image_path)
                image_filename = f"uploads/images/{image_filename}"
        
        # Handle document upload
        document_filename = ''
        if 'document_attachment' in request.files:
            document_file = request.files['document_attachment']
            if document_file and document_file.filename and allowed_file(document_file.filename, ALLOWED_DOCUMENT_EXTENSIONS):
                filename = secure_filename(document_file.filename)
                # Add asset code to filename for uniqueness
                name, ext = os.path.splitext(filename)
                document_filename = f"{asset_code}_{name}{ext}"
                document_path = os.path.join(DOCUMENT_FOLDER, document_filename)
                document_file.save(document_path)
                document_filename = f"uploads/documents/{document_filename}"
        
        asset_data = {
            'Asset Code': asset_code,
            'Item Name': item_name,
            'Asset Category': asset_category,
            'Asset SubCategory': asset_subcategory,
            'Brand': brand,
            'Asset Description': asset_description,
            'Amount': amount,
            'Location': location,
            'Date of Purchase': date_of_purchase,
            'Warranty': warranty,
            'Department': department,
            'Ownership': ownership,
            'Asset Status': request.form.get('asset_status', ''),
            'Image Attachment': image_filename,
            'Document Attachment': document_filename
        }
        
        if db.insert('Assets', asset_data):
            log_activity('Add', 'Asset', asset_code, f"Added new asset: {item_name}", f"Category: {asset_category}, Location: {location}")
            flash('Asset added successfully', 'success')
            return redirect(url_for('assets'))
        else:
            flash('Failed to add asset', 'danger')
    
    # Get master data for dropdowns
    categories = db.get_all('Categories')
    subcategories = db.get_all('Subcategories')
    brands = db.get_all('Brands')
    locations = db.get_all('Locations')
    
    return render_template('add_asset.html', categories=categories, 
                         subcategories=subcategories, brands=brands, locations=locations)

@app.route('/assets/edit/<asset_code>', methods=['GET', 'POST'])
@login_required
def edit_asset(asset_code):
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('assets'))
    
    asset = db.get_by_id('Assets', 'Asset Code', asset_code)
    
    if not asset:
        flash('Asset not found', 'danger')
        return redirect(url_for('assets'))
    
    if request.method == 'POST':
        # Get existing file paths
        image_filename = asset.get('Image Attachment', '')
        document_filename = asset.get('Document Attachment', '')
        
        # Handle image upload (only if new file is provided)
        if 'image_attachment' in request.files:
            image_file = request.files['image_attachment']
            if image_file and image_file.filename and allowed_file(image_file.filename, ALLOWED_IMAGE_EXTENSIONS):
                # Delete old image if exists
                if image_filename:
                    old_path = os.path.join(image_filename)
                    if os.path.exists(old_path):
                        try:
                            os.remove(old_path)
                        except:
                            pass
                
                # Save new image
                filename = secure_filename(image_file.filename)
                name, ext = os.path.splitext(filename)
                new_image_filename = f"{asset_code}_{name}{ext}"
                image_path = os.path.join(IMAGE_FOLDER, new_image_filename)
                image_file.save(image_path)
                image_filename = f"uploads/images/{new_image_filename}"
        
        # Handle document upload (only if new file is provided)
        if 'document_attachment' in request.files:
            document_file = request.files['document_attachment']
            if document_file and document_file.filename and allowed_file(document_file.filename, ALLOWED_DOCUMENT_EXTENSIONS):
                # Delete old document if exists
                if document_filename:
                    old_path = document_filename
                    if os.path.exists(old_path):
                        try:
                            os.remove(old_path)
                        except:
                            pass
                
                # Save new document
                filename = secure_filename(document_file.filename)
                name, ext = os.path.splitext(filename)
                new_document_filename = f"{asset_code}_{name}{ext}"
                document_path = os.path.join(DOCUMENT_FOLDER, new_document_filename)
                document_file.save(document_path)
                document_filename = f"uploads/documents/{new_document_filename}"
        
        asset_data = {
            'Item Name': request.form.get('item_name'),
            'Asset Category': request.form.get('asset_category'),
            'Asset SubCategory': request.form.get('asset_subcategory'),
            'Brand': request.form.get('brand'),
            'Asset Description': request.form.get('asset_description'),
            'Amount': request.form.get('amount'),
            'Location': request.form.get('location'),
            'Date of Purchase': request.form.get('date_of_purchase'),
            'Warranty': request.form.get('warranty'),
            'Department': request.form.get('department'),
            'Ownership': request.form.get('ownership'),
            'Asset Status': request.form.get('asset_status', ''),
            'Image Attachment': image_filename,
            'Document Attachment': document_filename
        }
        
        if db.update('Assets', 'Asset Code', asset_code, asset_data):
            log_activity('Update', 'Asset', asset_code, f"Updated asset: {asset.get('Item Name', asset_code)}", f"Category: {asset_data.get('Asset Category', '')}")
            flash('Asset updated successfully', 'success')
            return redirect(url_for('assets'))
        else:
            flash('Failed to update asset', 'danger')
    
    categories = db.get_all('Categories')
    subcategories = db.get_all('Subcategories')
    brands = db.get_all('Brands')
    locations = db.get_all('Locations')
    
    return render_template('edit_asset.html', asset=asset, categories=categories,
                         subcategories=subcategories, brands=brands, locations=locations)

@app.route('/assets/delete/<asset_code>', methods=['POST'])
@admin_required
def delete_asset(asset_code):
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('assets'))
    
    # Get asset before deleting to remove associated files
    asset = db.get_by_id('Assets', 'Asset Code', asset_code)
    if asset:
        # Delete image if exists
        image_path = asset.get('Image Attachment', '')
        if image_path and os.path.exists(image_path):
            try:
                os.remove(image_path)
            except:
                pass
        
        # Delete document if exists
        document_path = asset.get('Document Attachment', '')
        if document_path and os.path.exists(document_path):
            try:
                os.remove(document_path)
            except:
                pass
    
    if db.delete('Assets', 'Asset Code', asset_code):
        log_activity('Delete', 'Asset', asset_code, f"Deleted asset: {asset.get('Item Name', asset_code) if asset else asset_code}")
        flash('Asset deleted successfully', 'success')
    return redirect(url_for('assets'))

# Route to serve uploaded files
@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# Asset Movement Routes
@app.route('/asset_movements')
@login_required
def asset_movements():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    movements = db.get_all('AssetMovements')
    return render_template('asset_movements.html', movements=movements, role=session.get('role'))

@app.route('/asset_movements/add', methods=['GET', 'POST'])
@login_required
def add_asset_movement():
    if request.method == 'POST':
        asset_code = request.form.get('asset_code')
        from_location = request.form.get('from_location')
        to_location = request.form.get('to_location')
        notes = request.form.get('notes', '')
        
        movement_id = db.get_next_id('AssetMovements')
        movement_data = {
            'ID': movement_id,
            'Asset Code': asset_code,
            'From Location': from_location,
            'To Location': to_location,
            'Movement Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Moved By': session.get('user_id'),
            'Notes': notes
        }
        
        if db.insert('AssetMovements', movement_data):
            # Update asset location
            db.update('Assets', 'Asset Code', asset_code, {'Location': to_location})
            log_activity('Move', 'Asset', asset_code, f"Asset moved from {from_location} to {to_location}", notes)
            flash('Asset movement recorded successfully', 'success')
            return redirect(url_for('asset_movements'))
        else:
            flash('Failed to record movement', 'danger')
    
    assets = db.get_all('Assets')
    locations = db.get_all('Locations')
    return render_template('add_movement.html', assets=assets, locations=locations)

# Barcode Printing Routes
@app.route('/barcode/print', methods=['POST'])
@login_required
def print_barcodes():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('assets'))
        
    asset_codes = request.form.getlist('asset_codes')
    
    if not asset_codes:
        flash('No assets selected', 'warning')
        return redirect(url_for('assets'))
    
    # Get asset details
    assets_data = []
    for code in asset_codes:
        asset = db.get_by_id('Assets', 'Asset Code', code)
        if asset:
            assets_data.append(asset)
    
    if not assets_data:
        flash('No valid assets found', 'warning')
        return redirect(url_for('assets'))
    
    # Redirect to print preview page with asset codes
    codes_param = ','.join(asset_codes)
    return redirect(url_for('print_preview', codes=codes_param))

@app.route('/barcode/print-preview')
@login_required
def print_preview():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('assets'))
    
    asset_codes = request.args.get('codes', '').split(',')
    asset_codes = [code.strip() for code in asset_codes if code.strip()]
    
    # Get asset details
    assets_data = []
    for code in asset_codes:
        asset = db.get_by_id('Assets', 'Asset Code', code)
        if asset:
            assets_data.append(asset)
    
    return render_template('barcode_print.html', assets=assets_data)

@app.route('/barcode/preview')
@login_required
def preview_barcodes():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    asset_codes = request.args.getlist('codes')
    assets = []
    for code in asset_codes:
        asset = db.get_by_id('Assets', 'Asset Code', code)
        if asset:
            assets.append(asset)
    return render_template('barcode_preview.html', assets=assets)

# Reports Routes
@app.route('/reports/assets')
@login_required
def asset_report():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    
    # Get filter parameters
    category = request.args.get('category', '')
    location = request.args.get('location', '')
    department = request.args.get('department', '')
    search = request.args.get('search', '').lower()
    
    # Get all assets
    all_assets = db.get_all('Assets')
    
    # Apply filters
    filtered_assets = []
    for asset in all_assets:
        # Category filter
        if category and asset.get('Asset Category', '') != category:
            continue
        # Location filter
        if location and asset.get('Location', '') != location:
            continue
        # Department filter
        if department and asset.get('Department', '') != department:
            continue
        # Search filter
        if search:
            searchable_text = f"{asset.get('Asset Code', '')} {asset.get('Item Name', '')} {asset.get('Brand', '')}".lower()
            if search not in searchable_text:
                continue
        filtered_assets.append(asset)
    
    # Get master data for filters
    categories = db.get_all('Categories')
    locations = db.get_all('Locations')
    
    # Get unique departments
    departments = list(set([asset.get('Department', '') for asset in all_assets if asset.get('Department', '')]))
    departments.sort()
    
    return render_template('reports/asset_report.html', 
                         assets=filtered_assets, 
                         categories=categories,
                         locations=locations,
                         departments=departments,
                         role=session.get('role'),
                         current_category=category,
                         current_location=location,
                         current_department=department,
                         current_search=search)

@app.route('/reports/movements')
@login_required
def movement_report():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    
    # Get filter parameters
    asset_code = request.args.get('asset_code', '')
    from_location = request.args.get('from_location', '')
    to_location = request.args.get('to_location', '')
    moved_by = request.args.get('moved_by', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Get all movements
    all_movements = db.get_all('AssetMovements')
    
    # Apply filters
    filtered_movements = []
    for movement in all_movements:
        # Asset code filter
        if asset_code and movement.get('Asset Code', '') != asset_code:
            continue
        # From location filter
        if from_location and movement.get('From Location', '') != from_location:
            continue
        # To location filter
        if to_location and movement.get('To Location', '') != to_location:
            continue
        # Moved by filter
        if moved_by and movement.get('Moved By', '') != moved_by:
            continue
        # Date filters
        movement_date = movement.get('Movement Date', '')
        if date_from and movement_date < date_from:
            continue
        if date_to and movement_date > date_to:
            continue
        filtered_movements.append(movement)
    
    # Sort by date (newest first)
    filtered_movements.sort(key=lambda x: x.get('Movement Date', ''), reverse=True)
    
    # Get master data for filters
    assets = db.get_all('Assets')
    locations = db.get_all('Locations')
    
    # Get unique users
    users = list(set([movement.get('Moved By', '') for movement in all_movements if movement.get('Moved By', '')]))
    users.sort()
    
    return render_template('reports/movement_report.html',
                         movements=filtered_movements,
                         assets=assets,
                         locations=locations,
                         users=users,
                         role=session.get('role'),
                         current_asset_code=asset_code,
                         current_from_location=from_location,
                         current_to_location=to_location,
                         current_moved_by=moved_by,
                         current_date_from=date_from,
                         current_date_to=date_to)

@app.route('/logs')
@login_required
def logs():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    
    # Get filter parameters
    log_type = request.args.get('type', '')
    user = request.args.get('user', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Get all logs (combining movements and other activities)
    logs = []
    
    # Get activity logs from ActivityLogs sheet
    try:
        activity_logs = db.get_all('ActivityLogs')
        for log_entry in activity_logs:
            logs.append({
                'type': log_entry.get('Type', 'Activity'),
                'date': log_entry.get('Date & Time', ''),
                'user': log_entry.get('User', ''),
                'description': log_entry.get('Description', ''),
                'details': log_entry.get('Details', ''),
                'action': log_entry.get('Action', ''),
                'entity_type': log_entry.get('Entity Type', ''),
                'entity_id': log_entry.get('Entity ID', '')
            })
    except Exception as e:
        print(f"Error loading activity logs: {e}")
    
    # Get movements as logs
    movements = db.get_all('AssetMovements')
    for movement in movements:
        logs.append({
            'type': 'Movement',
            'date': movement.get('Movement Date', ''),
            'user': movement.get('Moved By', ''),
            'description': f"Asset {movement.get('Asset Code', '')} moved from {movement.get('From Location', '')} to {movement.get('To Location', '')}",
            'details': movement.get('Notes', ''),
            'action': 'Move',
            'entity_type': 'Asset',
            'entity_id': movement.get('Asset Code', '')
        })
    
    # Sort by date (newest first)
    logs.sort(key=lambda x: x.get('date', ''), reverse=True)
    
    # Apply filters
    filtered_logs = []
    for log in logs:
        # Type filter
        if log_type and log.get('type', '') != log_type:
            continue
        # User filter
        if user and log.get('user', '') != user:
            continue
        # Date filters
        log_date = log.get('date', '')
        if date_from and log_date < date_from:
            continue
        if date_to and log_date > date_to:
            continue
        filtered_logs.append(log)
    
    # Get unique users and types
    log_types = list(set([log.get('type', '') for log in logs]))
    log_types.sort()
    log_users = list(set([log.get('user', '') for log in logs if log.get('user', '')]))
    log_users.sort()
    
    return render_template('logs.html',
                         logs=filtered_logs,
                         log_types=log_types,
                         users=log_users,
                         role=session.get('role'),
                         current_type=log_type,
                         current_user=user,
                         current_date_from=date_from,
                         current_date_to=date_to)

# Excel Export Routes
@app.route('/reports/assets/export')
@login_required
def export_asset_report():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    
    # Get filter parameters (same as report)
    category = request.args.get('category', '')
    location = request.args.get('location', '')
    department = request.args.get('department', '')
    search = request.args.get('search', '').lower()
    
    # Get all assets
    all_assets = db.get_all('Assets')
    
    # Apply filters
    filtered_assets = []
    for asset in all_assets:
        if category and asset.get('Asset Category', '') != category:
            continue
        if location and asset.get('Location', '') != location:
            continue
        if department and asset.get('Department', '') != department:
            continue
        if search:
            searchable_text = f"{asset.get('Asset Code', '')} {asset.get('Item Name', '')} {asset.get('Brand', '')}".lower()
            if search not in searchable_text:
                continue
        filtered_assets.append(asset)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Report"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = ['Asset Code', 'Item Name', 'Category', 'Subcategory', 'Brand', 
               'Description', 'Amount', 'Location', 'Date of Purchase', 
               'Warranty', 'Department', 'Ownership']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for asset in filtered_assets:
        row = [
            asset.get('Asset Code', ''),
            asset.get('Item Name', ''),
            asset.get('Asset Category', ''),
            asset.get('Asset SubCategory', ''),
            asset.get('Brand', ''),
            asset.get('Asset Description', ''),
            asset.get('Amount', ''),
            asset.get('Location', ''),
            asset.get('Date of Purchase', ''),
            asset.get('Warranty', ''),
            asset.get('Department', ''),
            asset.get('Ownership', '')
        ]
        ws.append(row)
    
    # Apply borders and alignment to data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'asset_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=filename)

@app.route('/reports/movements/export')
@login_required
def export_movement_report():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    
    # Get filter parameters
    asset_code = request.args.get('asset_code', '')
    from_location = request.args.get('from_location', '')
    to_location = request.args.get('to_location', '')
    moved_by = request.args.get('moved_by', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Get all movements
    all_movements = db.get_all('AssetMovements')
    
    # Apply filters
    filtered_movements = []
    for movement in all_movements:
        if asset_code and movement.get('Asset Code', '') != asset_code:
            continue
        if from_location and movement.get('From Location', '') != from_location:
            continue
        if to_location and movement.get('To Location', '') != to_location:
            continue
        if moved_by and movement.get('Moved By', '') != moved_by:
            continue
        movement_date = movement.get('Movement Date', '')
        if date_from and movement_date < date_from:
            continue
        if date_to and movement_date > date_to:
            continue
        filtered_movements.append(movement)
    
    # Sort by date
    filtered_movements.sort(key=lambda x: x.get('Movement Date', ''), reverse=True)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Movement Report"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = ['ID', 'Date & Time', 'Asset Code', 'From Location', 'To Location', 
               'Moved By', 'Notes']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for movement in filtered_movements:
        row = [
            movement.get('ID', ''),
            movement.get('Movement Date', ''),
            movement.get('Asset Code', ''),
            movement.get('From Location', ''),
            movement.get('To Location', ''),
            movement.get('Moved By', ''),
            movement.get('Notes', '')
        ]
        ws.append(row)
    
    # Apply borders and alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'movement_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=filename)

@app.route('/depreciation')
@login_required
def depreciation():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    
    from datetime import datetime
    
    # Get filter parameters
    category = request.args.get('category', '')
    location = request.args.get('location', '')
    status = request.args.get('status', '')
    
    # Get all assets and asset types
    all_assets = db.get_all('Assets')
    asset_types = db.get_all('AssetTypes')
    categories = db.get_all('Categories')
    locations = db.get_all('Locations')
    
    # Create a lookup dictionary for asset types by Asset Type name
    asset_type_lookup = {}
    for at in asset_types:
        asset_type_name = at.get('Asset Type', '')
        dep_value = at.get('Depreciation Value (%)', '')
        try:
            asset_type_lookup[asset_type_name] = float(dep_value) if dep_value else 0
        except (ValueError, TypeError):
            asset_type_lookup[asset_type_name] = 0
    
    # Calculate depreciation for each asset
    today = datetime.now().date()
    processed_assets = []
    
    total_purchase_value = 0
    total_depreciation = 0
    total_current_value = 0
    
    for asset in all_assets:
        # Apply filters
        if category and asset.get('Asset Category', '') != category:
            continue
        if location and asset.get('Location', '') != location:
            continue
        if status and asset.get('Asset Status', '') != status:
            continue
        
        # Get purchase amount
        purchase_amount_str = asset.get('Amount', '').strip()
        purchase_amount = 0
        try:
            purchase_amount = float(purchase_amount_str) if purchase_amount_str else 0
        except (ValueError, TypeError):
            purchase_amount = 0
        
        # Get purchase date
        purchase_date_str = asset.get('Date of Purchase', '').strip()
        purchase_date = None
        age_years = 0
        
        if purchase_date_str:
            try:
                # Try different date formats
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y', '%m-%d-%Y']:
                    try:
                        purchase_date = datetime.strptime(purchase_date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                
                if purchase_date:
                    # Calculate age in years
                    delta = today - purchase_date
                    age_years = delta.days / 365.25
                    if age_years < 0:
                        age_years = 0
            except:
                pass
        
        # Get depreciation percentage
        # Try to match by Asset Type field in Assets (if it exists)
        # Otherwise, try to match by category or use a default
        asset_type_name = asset.get('Asset Type', '') or asset.get('Asset Category', '')
        depreciation_percent = asset_type_lookup.get(asset_type_name, 0)
        
        # Calculate depreciation
        annual_depreciation = 0
        total_dep = 0
        current_value = purchase_amount
        
        if purchase_amount > 0 and depreciation_percent > 0 and age_years > 0:
            annual_depreciation = purchase_amount * (depreciation_percent / 100)
            total_dep = annual_depreciation * age_years
            current_value = max(0, purchase_amount - total_dep)  # Can't go below 0
        
        # Add calculated fields to asset
        asset['purchase_amount'] = purchase_amount
        asset['depreciation_percent'] = depreciation_percent
        asset['age_years'] = age_years
        asset['annual_depreciation'] = annual_depreciation
        asset['total_depreciation'] = total_dep
        asset['current_value'] = current_value
        
        processed_assets.append(asset)
        
        # Update totals
        total_purchase_value += purchase_amount
        total_depreciation += total_dep
        total_current_value += current_value
    
    # Calculate filtered totals
    filtered_purchase_value = sum(a.get('purchase_amount', 0) for a in processed_assets)
    filtered_total_depreciation = sum(a.get('total_depreciation', 0) for a in processed_assets)
    filtered_current_value = sum(a.get('current_value', 0) for a in processed_assets)
    filtered_annual_depreciation = sum(a.get('annual_depreciation', 0) for a in processed_assets)
    
    return render_template('depreciation.html',
                         filtered_assets=processed_assets,
                         total_assets=len(all_assets),
                         total_purchase_value=total_purchase_value,
                         total_depreciation=total_depreciation,
                         total_current_value=total_current_value,
                         filtered_purchase_value=filtered_purchase_value,
                         filtered_total_depreciation=filtered_total_depreciation,
                         filtered_current_value=filtered_current_value,
                         filtered_annual_depreciation=filtered_annual_depreciation,
                         categories=categories,
                         locations=locations,
                         role=session.get('role'))

@app.route('/depreciation/export')
@login_required
def export_depreciation():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    
    from datetime import datetime
    
    # Get filter parameters
    category = request.args.get('category', '')
    location = request.args.get('location', '')
    status = request.args.get('status', '')
    
    # Get all assets and asset types
    all_assets = db.get_all('Assets')
    asset_types = db.get_all('AssetTypes')
    
    # Create lookup dictionary
    asset_type_lookup = {}
    for at in asset_types:
        asset_type_name = at.get('Asset Type', '')
        dep_value = at.get('Depreciation Value (%)', '')
        try:
            asset_type_lookup[asset_type_name] = float(dep_value) if dep_value else 0
        except (ValueError, TypeError):
            asset_type_lookup[asset_type_name] = 0
    
    # Calculate depreciation for each asset
    today = datetime.now().date()
    processed_assets = []
    
    for asset in all_assets:
        # Apply filters
        if category and asset.get('Asset Category', '') != category:
            continue
        if location and asset.get('Location', '') != location:
            continue
        if status and asset.get('Asset Status', '') != status:
            continue
        
        # Calculate values (same logic as main route)
        purchase_amount_str = asset.get('Amount', '').strip()
        purchase_amount = 0
        try:
            purchase_amount = float(purchase_amount_str) if purchase_amount_str else 0
        except (ValueError, TypeError):
            purchase_amount = 0
        
        purchase_date_str = asset.get('Date of Purchase', '').strip()
        purchase_date = None
        age_years = 0
        
        if purchase_date_str:
            try:
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y', '%m-%d-%Y']:
                    try:
                        purchase_date = datetime.strptime(purchase_date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                
                if purchase_date:
                    delta = today - purchase_date
                    age_years = delta.days / 365.25
                    if age_years < 0:
                        age_years = 0
            except:
                pass
        
        asset_type_name = asset.get('Asset Type', '') or asset.get('Asset Category', '')
        depreciation_percent = asset_type_lookup.get(asset_type_name, 0)
        
        annual_depreciation = 0
        total_dep = 0
        current_value = purchase_amount
        
        if purchase_amount > 0 and depreciation_percent > 0 and age_years > 0:
            annual_depreciation = purchase_amount * (depreciation_percent / 100)
            total_dep = annual_depreciation * age_years
            current_value = max(0, purchase_amount - total_dep)
        
        asset['purchase_amount'] = purchase_amount
        asset['depreciation_percent'] = depreciation_percent
        asset['age_years'] = age_years
        asset['annual_depreciation'] = annual_depreciation
        asset['total_depreciation'] = total_dep
        asset['current_value'] = current_value
        
        processed_assets.append(asset)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Depreciation Report"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = ['Asset Code', 'Item Name', 'Category', 'Location', 'Purchase Date', 
               'Age (Years)', 'Purchase Amount', 'Depreciation %', 'Annual Depreciation', 
               'Total Depreciation', 'Current Book Value', 'Status']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for asset in processed_assets:
        row = [
            asset.get('Asset Code', ''),
            asset.get('Item Name', ''),
            asset.get('Asset Category', ''),
            asset.get('Location', ''),
            asset.get('Date of Purchase', ''),
            round(asset.get('age_years', 0), 2) if asset.get('age_years') else '',
            round(asset.get('purchase_amount', 0), 2) if asset.get('purchase_amount') else '',
            f"{asset.get('depreciation_percent', 0)}%" if asset.get('depreciation_percent') else '',
            round(asset.get('annual_depreciation', 0), 2) if asset.get('annual_depreciation') else '',
            round(asset.get('total_depreciation', 0), 2) if asset.get('total_depreciation') else '',
            round(asset.get('current_value', 0), 2) if asset.get('current_value') is not None else '',
            asset.get('Asset Status', '')
        ]
        ws.append(row)
    
    # Add totals row
    if processed_assets:
        total_row = ['', '', '', '', '', 'TOTALS:',
                    sum(a.get('purchase_amount', 0) for a in processed_assets),
                    '',
                    sum(a.get('annual_depreciation', 0) for a in processed_assets),
                    sum(a.get('total_depreciation', 0) for a in processed_assets),
                    sum(a.get('current_value', 0) for a in processed_assets),
                    '']
        ws.append(total_row)
        
        # Style totals row
        for cell in ws[ws.max_row]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.border = border
    
    # Apply borders and alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', horizontal='left')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'depreciation_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=filename)

@app.route('/logs/export')
@login_required
def export_logs():
    if not db:
        flash('Database not configured', 'danger')
        return redirect(url_for('login'))
    
    # Get filter parameters
    log_type = request.args.get('type', '')
    user = request.args.get('user', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Get all logs
    logs = []
    movements = db.get_all('AssetMovements')
    for movement in movements:
        logs.append({
            'type': 'Movement',
            'date': movement.get('Movement Date', ''),
            'user': movement.get('Moved By', ''),
            'description': f"Asset {movement.get('Asset Code', '')} moved from {movement.get('From Location', '')} to {movement.get('To Location', '')}",
            'details': movement.get('Notes', '')
        })
    
    # Sort by date
    logs.sort(key=lambda x: x.get('date', ''), reverse=True)
    
    # Apply filters
    filtered_logs = []
    for log in logs:
        if log_type and log.get('type', '') != log_type:
            continue
        if user and log.get('user', '') != user:
            continue
        log_date = log.get('date', '')
        if date_from and log_date < date_from:
            continue
        if date_to and log_date > date_to:
            continue
        filtered_logs.append(log)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Logs"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = ['Date & Time', 'Type', 'User', 'Description', 'Details']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for log in filtered_logs:
        row = [
            log.get('date', ''),
            log.get('type', ''),
            log.get('user', ''),
            log.get('description', ''),
            log.get('details', '')
        ]
        ws.append(row)
    
    # Apply borders and alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'activity_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=filename)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(debug=debug, host='0.0.0.0', port=port)

